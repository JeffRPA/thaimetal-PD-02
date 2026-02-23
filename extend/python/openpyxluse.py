

###########################################################################
#openpyxluse.py
###########################################################################


def read_xlsx_with_cell_objects(file_path, sheet_name):
    import openpyxl
    # Load the workbook and select the specific sheet
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name]
    
    # Initialize a list to store the rows
    result = []
    
    # Iterate through each row in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Extract each cell value
        row_data = [cell if cell is not None else "" for cell in row]
        result.append(row_data)

    if workbook:
        workbook.close()  

    return result



def read_xlsx_with_cell_objects__get_only_not_hidden_rows(file_path, sheet_name):
    import openpyxl
    # Load the workbook and select the specific sheet without read_only mode to access row_dimensions
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    
    # Initialize a list to store the rows
    result = []
    
    # Get the header row to determine the number of columns
    header = next(sheet.iter_rows(values_only=True))
    num_columns = len(header)
    
    # Iterate through each row in the sheet, starting after the header
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):  # Retrieve only values (not cell objects)
        # Check if the row is hidden
        if not sheet.row_dimensions[idx].hidden:
            # Extract each cell value within the header range, replacing None with an empty string
            row_data = [row[i] if row[i] is not None else "" for i in range(num_columns)]
            result.append(row_data)

    workbook.close()

    return result



def numberofrows(listofarrays):
    return len(listofarrays)


def numberofcolumns(listofarrays):
    return len(listofarrays[len(listofarrays)-1])


def number_to_column_letter(n):
    column_letter = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        column_letter = chr(65 + remainder) + column_letter
    return column_letter

def lastcolumnletter(listofarrays):
    lastcolumnalphabetletter = number_to_column_letter(len(listofarrays[len(listofarrays)-1]))
    return str(lastcolumnalphabetletter)

def lastcolumnrow(listofarrays):
    lastcolumnalphabetletter = number_to_column_letter(len(listofarrays[len(listofarrays)-1]))
    lastrow = len(listofarrays)
    return str(lastcolumnalphabetletter) + str(lastrow)





def read_xlsm_as_list_of_arrays(file_path, sheet_name=None):

    from openpyxl import load_workbook

    """
    Reads data from an .xlsm file and returns it as a list of arrays.

    :param file_path: Path to the .xlsm file
    :param sheet_name: Name of the sheet to read (optional). If None, reads the active sheet.
    :return: List of rows, where each row is represented as a list of cell values
    """
    # Load the workbook
    workbook = load_workbook(file_path, keep_vba=True)
    
    # Select the sheet
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    
    # Read data into a list of arrays
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))  # Convert each row to a list
    
    return data





'''
# Original array with empty lists
array = [
    [5, 65, 8],
    [243, 54, 67],
    [],
    [],
    [],
    [],
    []
]
'''

# Remove empty lists using a list comprehension
def removetheemptylistsfromarray(input_array):
    filtered_array = [sublist for sublist in input_array if sublist]
    return filtered_array





def unhide_all_columns_xlsx_specficsheet(input_file_path, specific_sheetname):
    from openpyxl import load_workbook
    # Load the workbook
    wb = load_workbook(input_file_path)
    
    # specific sheet in the workbook
    ws = wb[specific_sheetname]
    # Iterate through all column dimensions
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.hidden:
            col_dim.hidden = False  # Unhide this column
    
    # Save the changes
    wb.save(input_file_path)




def unhide_all_columns_xlsx_all_sheets(input_file_path):
    from openpyxl import load_workbook
    # Load the workbook
    wb = load_workbook(input_file_path)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        # specific sheet in the workbook
        ws = wb[sheet_name]
        # Iterate through all column dimensions
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.hidden:
                col_dim.hidden = False  # Unhide this column
    
    # Save the changes
    wb.save(input_file_path)





def clear_except_first_row(file_path, sheet_name):
    import openpyxl
    """
    Clears all cell values in the specified sheet of the .xlsx file except those in the first row.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to modify.
    """
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]
    
    # Iterate over all rows starting from the second row and clear each cell's value
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None  # Clear the cell's value

    # Save the workbook (this overwrites the existing file)
    workbook.save(file_path)
    print(f"All cells except those in the first row have been cleared in '{sheet_name}'.")




def write_array_to_excel(file_path, data, sheet_name, start_cell):
    import openpyxl
    
    """
    Write a list of data to column A of an existing Excel file starting from a given cell.
    
    Args:
        file_path (str): The path to the existing Excel (.xlsx) file.
        data (list): The list of data to write into column A.
        sheet_name (str): The worksheet name where data will be written. Default is "Sheet1".
        start_cell (str): The starting cell (e.g., "A2") from where data will be written. Default is "A2".
    """
    # Load the existing workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Determine the starting row from start_cell (assumes format like 'A2')
    start_row = int(start_cell[1:])
    
    # Write each element of the data list to column A, starting from start_row
    for index, value in enumerate(data):
        ws.cell(row=start_row + index, column=1, value=value)
    
    # Save changes back to the Excel file
    wb.save(file_path)




def excel_to_dict(filename, sheet_name, key_column, value_column, start_row):
    import openpyxl
    from openpyxl.utils.cell import column_index_from_string
    """
    Reads an Excel sheet and returns a dict where:
      • keys   come from `key_column`
      • values come from `value_column`
    Reading starts at row `start_row` and goes down to the last non-empty row.
    
    Parameters:
      filename     – path to the .xlsx file
      sheet_name   – worksheet name
      key_column   – letter of the column to use as dict keys
      value_column – letter of the column to use as dict values
      start_row    – first row index to read (1-based)
      
    Returns:
      dict mapping each key to its corresponding value
    """
    wb = openpyxl.load_workbook(filename, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {filename!r}")
    ws = wb[sheet_name]

    # convert column letters to indices
    key_col_idx = column_index_from_string(key_column)
    val_col_idx = column_index_from_string(value_column)

    result = {}
    # iterate row-by-row to keep key/value paired correctly
    for row in range(start_row, ws.max_row + 1):
        key = ws.cell(row=row, column=key_col_idx).value
        val = ws.cell(row=row, column=val_col_idx).value
        
        # skip completely blank rows (optional)
        if key is None and val is None:
            continue
        
        result[key] = val

    return result

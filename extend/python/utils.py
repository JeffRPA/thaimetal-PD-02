
def array_to_array_string(arr) -> str:
    """
    Convert a Python list/tuple (including nested) into a string that looks
    like a Python array literal, e.g. [1, 2, ["a", "b"]].
    """
    import ast
    return repr(arr)

def reverse_array(a):
    b = list(reversed(a))
    return b

def remove_items_from_list(input_list, items_to_remove):
    return [x for x in input_list if x not in items_to_remove]

def find_by_second_index(data, index, target):
    result = []
    for row in data:
        if len(row) > 1 and row[index] == target:
            result.append(row)
    return result



def remove_processed_rows(data_2d, updated_rows, po_index=2):
    import json
    try:
        updated_rows = json.loads(updated_rows)
    except Exception as e:
        print("JSON parse error:", e)
        return data_2d
    processed_pos = set()
    for item in updated_rows:
        if isinstance(item, dict) and 'po' in item:
            processed_pos.add(item['po'])
    data_2d[:] = [
        row for row in data_2d
        if len(row) > po_index and row[po_index] not in processed_pos
    ]
    return data_2d

def get_list_of_arrays_firstrowheader_and_secondrowonwards(xlsx_filepath, sheetidentifier, sheetidentifiedtype):
    import pandas as pd
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_filepath, data_only=True)

    if sheetidentifiedtype == "name":
        ws = wb[sheetidentifier]
    elif sheetidentifiedtype == "index":
        ws = wb.worksheets[sheetidentifier]
    else:
        raise ValueError("sheetidentifiedtype ???????? 'name' ???? 'index'")

    rows = list(ws.iter_rows(values_only=True))
    raw_df = pd.DataFrame(rows)

    header = raw_df.iloc[0].tolist()
    df = raw_df.iloc[1:].reset_index(drop=True)
    df.columns = header

    df = df.dropna(how="all").reset_index(drop=True)

    list_of_arrays = [header] + [row.tolist() for _, row in df.iterrows()]
    return list_of_arrays